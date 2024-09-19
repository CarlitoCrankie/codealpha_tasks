import speech_recognition as sr
import pyttsx3
import pywhatkit
import datetime
import wikipedia
import requests
from openpyxl import load_workbook
import random
import time
import threading
import subprocess
import webbrowser
import jarvisGUI  # Import the GUI module

# Initialize recognizer and text-to-speech engine
listener = sr.Recognizer()
machine = pyttsx3.init()

# Global variables for caching
command_list, reply_list = [], []
weather_cache = {}
last_weather_update = None

# Function to make the assistant speak
def talk(text):
    machine.say(text)
    machine.runAndWait()

# Function to listen for user instructions with timeout
def input_instruction():
    global instruction
    try:
        with sr.Microphone() as origin:
            print("Listening")
            speech = listener.listen(origin, timeout=5, phrase_time_limit=10)  # Set timeout
            instruction = listener.recognize_google(speech)
            instruction = instruction.lower()
            if 'jarvis' in instruction:
                instruction = instruction.replace('jarvis', "").strip()
                print("Processed Instruction:", instruction)
    except sr.UnknownValueError:
        print("Sorry, I did not understand that.")
        return ""
    except sr.RequestError:
        print("Sorry, there was an error with the speech recognition service.")
        return ""
    except Exception as e:
        print(f"An error occurred: {e}")
        return ""
    return instruction

# Weather-related functions
def get_temperature(json_data):
    return json_data['main']['temp']

def get_weather_type(json_data):
    return json_data['weather'][0]['description']

def get_wind_speed(json_data):
    return json_data['wind']['speed']

def get_weather_data(json_data, city):
    weather_type = get_weather_type(json_data)
    temperature = get_temperature(json_data)
    wind_speed = get_wind_speed(json_data)
    return f"The weather in {city} is currently {weather_type} with a temperature of {temperature} degrees and wind speeds reaching {wind_speed} kilometers per hour."

def weather(city='Takoradi'):
    global weather_cache, last_weather_update
    if last_weather_update and time.time() - last_weather_update < 600:  # Cache for 10 minutes
        return weather_cache.get(city)
    
    api_key = '89a12bb697c434fb9a5d23a5d8c638b3'  # Replace with your real API key
    api_address = f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric'
    json_data = requests.get(api_address).json()
    
    if "weather" in json_data:
        weather_data = get_weather_data(json_data, city)
        weather_cache[city] = weather_data
        last_weather_update = time.time()
        return weather_data
    else:
        return "Sorry, I couldn't retrieve the weather information right now."

# Load Excel data only once
def excel():
    global command_list, reply_list
    if not command_list and not reply_list:  # Only load once
        wb = load_workbook("Data/input.xlsx")
        wu = wb['User']
        wr = wb['Replies']

        command_list = [[cell.strip().lower() for cell in row if cell] for row in wu.iter_rows(min_row=1, max_col=wu.max_column, values_only=True)]
        reply_list = [[cell.strip() for cell in row if cell] for row in wr.iter_rows(min_row=1, max_col=wr.max_column, values_only=True)]

        print("Command List:", command_list)
        print("Reply List:", reply_list)

# Function to play Jarvis' actions
def play_Jarvis():
    instruction = input_instruction()
    print("Received Instruction:", instruction)

    if "play" in instruction:
        song = instruction.replace('play', "").strip()
        talk("Playing " + song)
        thread = threading.Thread(target=pywhatkit.playonyt, args=(song,))  # Non-blocking
        thread.start()

    elif 'time' in instruction:
        current_time = datetime.datetime.now().strftime('%I:%M %p')
        talk('Current time is ' + current_time)

    elif 'date' in instruction:
        current_date = datetime.datetime.now().strftime('%d/%m/%Y')
        talk('Today\'s date is ' + current_date)

    elif 'open' in instruction:
        x = lambda: webbrowser.open_new('http://google.com')
        t = threading.Thread(target=x)  # Run in a separate thread to avoid blocking
        t.start()

    elif 'start' in instruction:
        subprocess.Popen('C:\\Program Files\\Arduino IDE\\Arduino IDE.exe')  # Non-blocking subprocess

    elif 'weather' in instruction:
        weather_details = weather()
        talk(weather_details)

    elif 'launch gui' in instruction:
        talk("Launching GUI")
        gui_thread = threading.Thread(target=jarvisGUI.face)
        gui_thread.start()

    elif 'stop' in instruction or 'quit' in instruction:
        talk("Stopping GUI")
        jarvisGUI.stop_face()  # Call the stop function to end the GUI loop
        return

    else:
        for i, commands in enumerate(command_list):
            if any(command in instruction for command in commands):
                talk(random.choice(reply_list[i]))
                return
        talk("I didn't quite catch that. Could you please repeat?")

# Main loop
excel()
while True:
    play_Jarvis()